from openpyxl import Workbook
from openpyxl.styles import Font, fills, Alignment, PatternFill
from datetime import datetime
from django.conf import settings
import os

font_10bold = Font(name='Calibri', size=10, bold=True)
font_10 = Font(name='Calibri', size=10)

class QueryReport():
    def __init__(self):
        self.wb = Workbook()
        self._first_sheet = True
        self._filename = None

    def _getNewSheet(self):
        if self._first_sheet:
            self._first_sheet = False
            return self.wb.active
        else:
            return self.wb.create_sheet()

    def generateReportSheet(self, title, headers, data):
        ws = self._getNewSheet()
        ws.append([title])
        ws.rows[0][0].font = font_10bold
        for i,h in enumerate(headers):
            ws.append(h)
            ws.row_dimensions[2+i].font = font_10bold
        for j,d in enumerate(data):
            ws.append(d)
            ws.row_dimensions[3+j+i].font = font_10

    def save(self, username):
        name = settings.MEDIA_ROOT + 'reports/' + username + '_' + str(datetime.now()) + '.xlsx'
        self.wb.save(name)
        self._filename = name
        return name

    def remove(self):
        os.remove(self._filename)


