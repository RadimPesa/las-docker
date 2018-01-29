#!/usr/bin/python
# Set up the Django Environment
import sys, argparse
sys.path.append('/srv/www/biobank')
from django.core.management import setup_environ 
import settings
setup_environ(settings)
from catissue.tissue.models import *
from catissue.tissue.utils import *
from catissue.api.handlers import *
from django.http import HttpRequest
from openpyxl.reader.excel import load_workbook

#serve per scandire il file con il riepilogo delle piastre diatech per vedere se esistono quei campioni
def ReadFile(args):
    try:
        enable_graph()
        disable_graph()
        fin=args.fin
        f = open(fin, 'rb')
        wb = load_workbook(f, use_iterators = True)
        i=0
        lisgentot=[]
        #chiave il foglio e valore un dizionario con chiave il gen e valore l'indice all'interno del foglio
        dizgen={}
        for ws in wb:
            foglio = ws.title
            #print 'foglio',foglio
            diztemp={}
            j=2
            for row in ws.iter_rows('A2:C110'):                
                gen=row[0].value
                if gen!='WATER' and gen!='H2O' and gen!='H20':
                    if row[2].value!=None:
                        gen=row[2].value
                    #print 'gen',gen
                    lisgentot.append(gen)
                    diztemp[gen]=j
                    j+=1
            dizgen[foglio]=diztemp
            i+=1
            #if i==1:
                #break
        lisaliq=Aliquot.objects.filter(uniqueGenealogyID__in=lisgentot,availability=1).values_list('uniqueGenealogyID',flat=True)
        #print 'lisaliq',lisaliq
        k=0
        for foglio,diztemp in dizgen.items():
            for gen, indice in diztemp.items():
                if gen not in lisaliq:
                    #print 'gen',gen
                    #print 'foglio',foglio
                    #print 'indice',indice
                    k+=1
                    #print 'foglio: '+foglio+' gen: '+gen+' indice: '+str(indice)
                    print gen
        print 'k',k
        #print 'listot',lisgentot
        #print 'diztot',dizgen
    except Exception,e:
        print 'err',e
    return

if __name__=='__main__':
    parser = argparse.ArgumentParser(description='Check integrity data')
    parser.add_argument('--fin', help='Excel input file with data')
    args = parser.parse_args()
    ReadFile(args)
    
